#!/usr/bin/env python3
"""Extract a monthly TSLA accumulation series from Networth Tracker.xlsx.

The source workbook contains a purchase-only view in the
`Theoretical Transaction History` sheet. This script converts that ledger into
one row per month with cumulative split-adjusted shares and monthly USD spend.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path


@dataclass
class Txn:
    trade_date: date
    qty: float
    buy_usd: float


def month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def next_month(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)


def to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_trade_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def load_rows(workbook_path: Path, sheet_name: str, symbol: str) -> tuple[list[Txn], float | None]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Install openpyxl first: python3 -m pip install openpyxl") from exc

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration as exc:
        raise SystemExit(f"Sheet is empty: {sheet_name}") from exc

    header_index = {str(value).strip(): idx for idx, value in enumerate(headers) if value not in (None, "")}

    required = ["Date", "Transaction Type", "Symbol", "Qty", "Qty adjusted", "Total buy price (USD)"]
    missing = [name for name in required if name not in header_index]
    if missing:
        raise SystemExit(f"Missing required columns in {sheet_name}: {', '.join(missing)}")

    reference_price = None
    for idx, value in enumerate(headers):
        if value == "<-- Current price of TSLA" and idx > 0:
            reference_price = to_float(headers[idx - 1])
            break

    extracted: list[Txn] = []
    for row in rows:
        trade_type = row[header_index["Transaction Type"]]
        row_symbol = row[header_index["Symbol"]]
        if trade_type != "Buy" or row_symbol != symbol:
            continue

        trade_date = parse_trade_date(row[header_index["Date"]])
        if trade_date is None:
            continue

        qty = to_float(row[header_index["Qty adjusted"]])
        if qty is None:
            qty = to_float(row[header_index["Qty"]])
        buy_usd = to_float(row[header_index["Total buy price (USD)"]])
        if qty is None or buy_usd is None:
            continue

        extracted.append(Txn(trade_date=trade_date, qty=qty, buy_usd=buy_usd))

    if not extracted:
        raise SystemExit(f"No {symbol} buy rows found in {sheet_name}.")

    extracted.sort(key=lambda txn: txn.trade_date)
    return extracted, reference_price


def build_payload(
    rows: list[Txn],
    baseline_month: date,
    source_sheet: str,
    symbol: str,
    reference_price: float | None,
) -> dict[str, object]:
    monthly: dict[date, dict[str, float]] = defaultdict(lambda: {"net_shares": 0.0, "buy_usd": 0.0})
    for row in rows:
        month = month_start(row.trade_date)
        monthly[month]["net_shares"] += row.qty
        monthly[month]["buy_usd"] += row.buy_usd

    normalized_baseline = month_start(baseline_month)
    first_trade = rows[0].trade_date
    last_trade = rows[-1].trade_date
    current_month = normalized_baseline
    final_month = month_start(last_trade)
    cumulative_shares = 0.0
    cumulative_buy_usd = 0.0
    points: list[dict[str, object]] = []
    while current_month <= final_month:
        entry = monthly[current_month]
        cumulative_shares += entry["net_shares"]
        cumulative_buy_usd += entry["buy_usd"]
        points.append(
            {
                "month": current_month.isoformat(),
                "net_shares": round(entry["net_shares"], 4),
                "buy_usd": round(entry["buy_usd"], 4),
                "cumulative_shares": round(cumulative_shares, 4),
                "cumulative_buy_usd": round(cumulative_buy_usd, 4),
            }
        )
        current_month = next_month(current_month)

    return {
        "meta": {
            "symbol": symbol,
            "source_sheet": source_sheet,
            "series_type": "purchase_only_accumulation",
            "baseline_start_month": normalized_baseline.isoformat(),
            "first_recorded_trade_date": first_trade.isoformat(),
            "last_recorded_trade_date": last_trade.isoformat(),
            "reference_price_usd": reference_price,
            "purchase_count": len(rows),
            "final_cumulative_shares": round(cumulative_shares, 4),
            "final_cumulative_buy_usd": round(cumulative_buy_usd, 4),
        },
        "points": points,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to Networth Tracker.xlsx")
    parser.add_argument("--output", type=Path, required=True, help="Output JSON path")
    parser.add_argument("--sheet", default="Theoretical Transaction History", help="Sheet to extract from")
    parser.add_argument("--symbol", default="TSLA", help="Ticker symbol to filter")
    parser.add_argument(
        "--baseline-month",
        default="2019-11-01",
        help="Month to begin the published series (YYYY-MM-DD).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        baseline_month = datetime.strptime(args.baseline_month, "%Y-%m-%d").date()
    except ValueError as exc:
        raise SystemExit("--baseline-month must be YYYY-MM-DD") from exc

    rows, reference_price = load_rows(args.workbook, args.sheet, args.symbol)
    payload = build_payload(rows, baseline_month, args.sheet, args.symbol, reference_price)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(
        f"Wrote {len(payload['points'])} monthly points "
        f"({payload['meta']['first_recorded_trade_date']} to {payload['meta']['last_recorded_trade_date']}) "
        f"to {args.output}"
    )


if __name__ == "__main__":
    main()
