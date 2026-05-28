#!/usr/bin/env python3
"""Extract simple net-worth chart data from an XLSX workbook.

The script looks for a worksheet with Date and Networth/Net Worth columns and
writes JSON compatible with conviction.html's chart loader.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path


def normalize(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def find_columns(headers: list[object]) -> tuple[int, int]:
    normalized = [normalize(header) for header in headers]
    date_idx = next((i for i, name in enumerate(normalized) if name == "date"), -1)
    worth_idx = next(
        (i for i, name in enumerate(normalized) if name in {"networth", "networthusd", "totalnetworth"}),
        -1,
    )
    return date_idx, worth_idx


def extract_rows(workbook_path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Install openpyxl first: python3 -m pip install openpyxl") from exc

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            continue

        date_idx, worth_idx = find_columns(headers)
        if date_idx < 0 or worth_idx < 0:
            continue

        extracted: list[dict[str, object]] = []
        for row in rows:
            date_value = row[date_idx] if date_idx < len(row) else None
            worth_value = row[worth_idx] if worth_idx < len(row) else None
            if not date_value or worth_value in (None, ""):
                continue
            extracted.append({"date": str(date_value), "networth": float(worth_value)})

        if extracted:
            return extracted

    raise SystemExit("No sheet with Date and Networth columns was found.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to Networth Tracker.xlsx")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/networth_data.json"),
        help="Output JSON path, default: data/networth_data.json",
    )
    args = parser.parse_args()

    data = extract_rows(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"Wrote {len(data)} rows to {args.output}")


if __name__ == "__main__":
    main()
